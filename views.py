from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QSplitter, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QComboBox, QCheckBox, QTextEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QScrollArea, QFrame,
    QDialog, QMessageBox, QGridLayout, QAbstractItemView, QStackedWidget,
    QTabWidget, QFileDialog,
)
from PyQt6.QtCore import Qt, pyqtSignal, QSettings
from PyQt6.QtGui import QColor, QShortcut, QKeySequence

import os

from constants import APP_TITLE, C, DB_PATH, AREA_DISP, AREA_MAP, RELATIONS


# ── Widget helpers ────────────────────────────────────────────────────────────

def fv(row, k):
    try:
        v = row[k]
        return str(v).strip() if v else ""
    except Exception:
        return ""

def mk_btn(text, oid="btn_accent"):
    b = QPushButton(text)
    b.setObjectName(oid)
    b.setCursor(Qt.CursorShape.PointingHandCursor)
    return b

def shdr(text):
    l = QLabel(text)
    l.setObjectName("section_hdr")
    return l

def vbox_field(label, widget, bg=None):
    w = QWidget()
    if bg:
        w.setStyleSheet(f"background:{bg};")
    vb = QVBoxLayout(w)
    vb.setContentsMargins(0, 0, 0, 0)
    vb.setSpacing(2)
    lb = QLabel(label)
    lb.setObjectName("fl")
    vb.addWidget(lb)
    vb.addWidget(widget)
    return w

def mk_entry(val=""):
    e = QLineEdit()
    e.setText(str(val).strip() if val else "")
    return e

def mk_combo(values, val=""):
    cb = QComboBox()
    cb.addItems(values)
    if val:
        i = cb.findText(str(val).strip(), Qt.MatchFlag.MatchContains)
        if i >= 0:
            cb.setCurrentIndex(i)
    return cb

def mk_check(label, val=""):
    cb = QCheckBox(label)
    cb.setChecked(str(val).strip() == "Y")
    return cb

def ge(w):
    if isinstance(w, QLineEdit):
        return w.text().strip()
    if isinstance(w, QComboBox):
        return w.currentText().strip()
    if isinstance(w, QTextEdit):
        return w.toPlainText().strip()
    return ""


# ── Parishioner Form ──────────────────────────────────────────────────────────

class ParishionerForm(QDialog):
    def __init__(self, parent, db, pno=None, on_save=None, prefill_host="", prefill_area=""):
        super().__init__(parent)
        self.db = db
        self.pno = pno
        self.on_save = on_save
        self.setWindowTitle("교적 추가" if not pno else "교적 수정")
        self.resize(860, 700)
        self.setModal(True)

        existing = db.get(pno) if pno else None
        ev = lambda k: fv(existing, k) if existing else ""

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        body = QWidget()
        body.setObjectName("card")
        scroll.setWidget(body)
        grid = QGridLayout(body)
        grid.setContentsMargins(16, 12, 16, 12)
        grid.setHorizontalSpacing(16)
        grid.setVerticalSpacing(6)
        for col in range(4):
            grid.setColumnStretch(col, 1)
        outer.addWidget(scroll, 1)

        r = 0

        def add(label, widget, row, col, span=1):
            grid.addWidget(vbox_field(label, widget, C['card']), row, col, 1, span)
            return widget

        grid.addWidget(shdr("📋  기본 정보"), r, 0, 1, 4); r += 1

        if existing:
            cur_area = next((d for d in AREA_DISP if d.startswith(ev("parishioner_no")[:5])), AREA_DISP[-1])
        elif prefill_area:
            cur_area = next((d for d in AREA_DISP if d.startswith(prefill_area)), AREA_DISP[-1])
        else:
            cur_area = AREA_DISP[-1]

        self.area_cb = mk_combo(AREA_DISP, cur_area)
        self.no_le = QLineEdit(ev("parishioner_no") if existing else "")
        self.no_le.setReadOnly(True)
        self.no_le.setStyleSheet(f"background:{C['header']};color:{C['muted']};")
        grid.addWidget(vbox_field("구역 *", self.area_cb, C['card']), r, 0, 1, 2)
        grid.addWidget(vbox_field("교적번호 (자동 생성)", self.no_le, C['card']), r, 2, 1, 2); r += 1

        self.name_e  = add("이름 *",       mk_entry(ev("name")),                                        r, 0)
        self.host_e  = add("세대주 이름 *", mk_entry(ev("host_nm") if existing else prefill_host),      r, 1)
        self.rel_cb  = add("관계 *",        mk_combo(RELATIONS, ev("relation") if existing else "본인"), r, 2)
        self.bname_e = add("세례명",        mk_entry(ev("baptism_nm")),                                  r, 3); r += 1
        self.bday_e  = add("축일 (MM/DD)",      mk_entry(ev("baptism_day")),    r, 0)
        self.pid_e   = add("주민번호",           mk_entry(ev("personal_no")),   r, 1)
        self.reg_e   = add("등록일 (YYYY/MM/DD)", mk_entry(ev("registion_date")), r, 2); r += 1

        grid.addWidget(shdr("상태 플래그"), r, 0, 1, 4); r += 1
        self.etemal_cb = mk_check("영원한 교적", ev("etemal"))
        self.duty_cb   = mk_check("의무금 납부",  ev("money_duty_flag"))
        self.lazy_cb   = mk_check("냉담자",       ev("lazy_flag"))
        self.alone_cb  = mk_check("독거",         ev("alone_flag"))
        fw = QWidget(); fw.setObjectName("card")
        fl = QHBoxLayout(fw); fl.setContentsMargins(4, 2, 4, 2)
        for cb in [self.etemal_cb, self.duty_cb, self.lazy_cb, self.alone_cb]:
            fl.addWidget(cb)
        fl.addStretch()
        grid.addWidget(fw, r, 0, 1, 4); r += 1

        grid.addWidget(shdr("📞  연락처 & 주소"), r, 0, 1, 4); r += 1
        self.telh_e  = add("집 전화",   mk_entry(ev("tel_home")),   r, 0)
        self.telm_e  = add("휴대폰",    mk_entry(ev("tel_hp")),     r, 1)
        self.telo_e  = add("직장 전화", mk_entry(ev("tel_office")), r, 2); r += 1
        self.addr_e  = add("주소",      mk_entry(ev("address")),    r, 0, 2)
        self.addrn_e = add("상세 주소", mk_entry(ev("address_no")), r, 2, 2); r += 1

        grid.addWidget(shdr("✝  세례 정보"), r, 0, 1, 4); r += 1
        self.bp_nm_e = add("세례 교구",           mk_entry(ev("baptism_parish_nm")), r, 0)
        self.bp_ch_e = add("세례 성당",           mk_entry(ev("baptism_church")),    r, 1)
        self.bp_dt_e = add("세례일 (YYYY/MM/DD)", mk_entry(ev("baptism_date")),      r, 2)
        self.bp_no_e = add("세례 번호",           mk_entry(ev("baptism_no")),        r, 3); r += 1

        grid.addWidget(shdr("🕊  견진 정보"), r, 0, 1, 4); r += 1
        self.sp_nm_e = add("견진 교구",           mk_entry(ev("sacrament_parish_nm")), r, 0)
        self.sp_ch_e = add("견진 성당",           mk_entry(ev("sacrament_church")),    r, 1)
        self.sp_dt_e = add("견진일 (YYYY/MM/DD)", mk_entry(ev("sacrament_date")),      r, 2); r += 1

        grid.addWidget(shdr("💼  직장 / 이전 소속"), r, 0, 1, 4); r += 1
        self.ofnm_e   = add("직장명",    mk_entry(ev("office_nm")),         r, 0)
        self.jknd_e   = add("직종",      mk_entry(ev("job_kind")),          r, 1)
        self.pre_nm_e = add("이전 교구", mk_entry(ev("pre_parish_nm")),     r, 2)
        self.pre_ch_e = add("이전 성당", mk_entry(ev("pre_parish_church")), r, 3); r += 1

        grid.addWidget(shdr("📝  메모"), r, 0, 1, 4); r += 1
        self.memo_te = QTextEdit()
        self.memo_te.setFixedHeight(80)
        self.memo_te.setPlainText(ev("personal_memo"))
        grid.addWidget(vbox_field("메모", self.memo_te, C['card']), r, 0, 1, 4); r += 1
        grid.setRowStretch(r, 1)

        bb = QWidget(); bb.setObjectName("card"); bb.setFixedHeight(54)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(12, 8, 12, 8)
        bbl.addStretch()
        cancel_b = mk_btn("취소", "btn_muted")
        save_b   = mk_btn("💾  저장", "btn_accent")
        cancel_b.clicked.connect(self.reject)
        save_b.clicked.connect(self._save)
        bbl.addWidget(cancel_b); bbl.addWidget(save_b)
        outer.addWidget(bb)

    def _save(self):
        name = ge(self.name_e)
        host = ge(self.host_e)
        if not name:
            QMessageBox.warning(self, "오류", "이름을 입력하세요.")
            return
        if not host:
            QMessageBox.warning(self, "오류", "세대주 이름을 입력하세요.")
            return
        area_code = ge(self.area_cb).split()[0]
        pno = self.pno if self.pno else self.db.next_no(area_code)
        data = dict(
            parishioner_no=pno, name=name, host_nm=host,
            relation=ge(self.rel_cb), baptism_nm=ge(self.bname_e),
            baptism_day=ge(self.bday_e), personal_no=ge(self.pid_e),
            registion_date=ge(self.reg_e),
            etemal="Y" if self.etemal_cb.isChecked() else "N",
            money_duty_flag="Y" if self.duty_cb.isChecked() else "N",
            lazy_flag="Y" if self.lazy_cb.isChecked() else "N",
            alone_flag="Y" if self.alone_cb.isChecked() else "N",
            tel_home=ge(self.telh_e), tel_hp=ge(self.telm_e), tel_office=ge(self.telo_e),
            address=ge(self.addr_e), address_no=ge(self.addrn_e),
            baptism_parish_nm=ge(self.bp_nm_e), baptism_church=ge(self.bp_ch_e),
            baptism_date=ge(self.bp_dt_e), baptism_no=ge(self.bp_no_e),
            sacrament_parish_nm=ge(self.sp_nm_e), sacrament_church=ge(self.sp_ch_e),
            sacrament_date=ge(self.sp_dt_e),
            office_nm=ge(self.ofnm_e), job_kind=ge(self.jknd_e),
            pre_parish_nm=ge(self.pre_nm_e), pre_parish_church=ge(self.pre_ch_e),
            personal_memo=ge(self.memo_te),
        )
        try:
            if self.pno:
                del data["parishioner_no"]
                self.db.update(self.pno, data)
            else:
                self.db.create(data)
            if self.on_save:
                self.on_save()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", str(e))


# ── Detail Panel ──────────────────────────────────────────────────────────────

class DetailPanel(QWidget):
    edit_sig        = pyqtSignal(str)
    delete_sig      = pyqtSignal(str, bool)
    add_sig         = pyqtSignal(str, str)
    perm_delete_sig = pyqtSignal(str)

    def __init__(self, db):
        super().__init__()
        self.db = db
        self._l = QVBoxLayout(self)
        self._l.setContentsMargins(0, 0, 0, 0)
        self._l.setSpacing(0)
        self._current_tab = 0
        self.show_empty()

    def _clear(self):
        while self._l.count():
            it = self._l.takeAt(0)
            if it.widget():
                it.widget().deleteLater()

    def show_empty(self):
        self._clear()
        w = QWidget(); w.setObjectName("card")
        l = QVBoxLayout(w)
        lbl = QLabel("← 교적을 선택하세요")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet(f"color:{C['muted']};font-size:15px;")
        l.addWidget(lbl)
        self._l.addWidget(w)

    def load(self, pno):
        p = self.db.get(pno)
        if not p:
            self.show_empty()
            return
        self._clear()

        v = lambda k: fv(p, k) or "—"
        is_del = v("delete_flag") == "Y"

        hdr = QWidget(); hdr.setObjectName("detail_hdr"); hdr.setFixedHeight(68)
        hl = QHBoxLayout(hdr); hl.setContentsMargins(16, 8, 16, 8)
        name_txt = f"{v('name')}  ({v('baptism_nm')})" if v("baptism_nm") != "—" else v("name")
        nl = QLabel(name_txt)
        nl.setStyleSheet("font-size:17px;font-weight:bold;")
        sl = QLabel(f"교적번호: {v('parishioner_no')}  |  관계: {v('relation')}  |  세대주: {v('host_nm')}")
        sl.setStyleSheet("color:#BDD7EE;font-size:10px;")
        close_btn = QPushButton("✕")
        close_btn.setFixedSize(24, 24)
        close_btn.setStyleSheet(
            "QPushButton{background:transparent;color:white;font-size:14px;"
            "border:none;border-radius:4px;padding:0;}"
            "QPushButton:hover{background:rgba(255,255,255,0.2);}"
        )
        close_btn.clicked.connect(self.show_empty)
        info_col = QVBoxLayout(); info_col.addWidget(nl); info_col.addWidget(sl)
        hl.addLayout(info_col); hl.addStretch(); hl.addWidget(close_btn, 0, Qt.AlignmentFlag.AlignTop)
        self._l.addWidget(hdr)

        bb = QWidget(); bb.setObjectName("detail_btnbar"); bb.setFixedHeight(46)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(8, 6, 8, 6); bbl.setSpacing(6)
        eb     = mk_btn("✏️  수정", "btn_accent")
        ab     = mk_btn("👤+  가족 추가", "btn_success")
        db_btn = mk_btn("♻️  복원" if is_del else "🗑  삭제",
                        "btn_accent" if is_del else "btn_danger")
        eb.clicked.connect(lambda: self.edit_sig.emit(pno))
        ab.clicked.connect(lambda: self.add_sig.emit(v("host_nm"), pno))
        db_btn.clicked.connect(lambda: self.delete_sig.emit(pno, is_del))
        bbl.addWidget(eb); bbl.addWidget(ab); bbl.addWidget(db_btn)
        if is_del:
            pd_btn = mk_btn("🗑  영구 삭제", "btn_danger")
            pd_btn.clicked.connect(lambda: self.perm_delete_sig.emit(pno))
            bbl.addWidget(pd_btn)
        bbl.addStretch()
        self._l.addWidget(bb)

        sc = QScrollArea(); sc.setWidgetResizable(True); sc.setFrameShape(QFrame.Shape.NoFrame)
        body = QWidget(); body.setObjectName("card")
        grid = QGridLayout(body)
        grid.setContentsMargins(12, 8, 12, 16)
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(4)
        grid.setColumnStretch(1, 1); grid.setColumnStretch(3, 1)
        sc.setWidget(body)

        r = 0

        def sh(t):
            nonlocal r
            grid.addWidget(shdr(t), r, 0, 1, 4); r += 1

        def r2(l1, v1, l2="", v2=""):
            nonlocal r
            lw = QLabel(l1); lw.setObjectName("fl")
            vw = QLabel(v1); vw.setObjectName("fv"); vw.setWordWrap(True)
            grid.addWidget(lw, r, 0); grid.addWidget(vw, r, 1)
            if l2:
                lw2 = QLabel(l2); lw2.setObjectName("fl")
                vw2 = QLabel(v2); vw2.setObjectName("fv"); vw2.setWordWrap(True)
                grid.addWidget(lw2, r, 2); grid.addWidget(vw2, r, 3)
            r += 1

        sh("📋  기본 정보")
        r2("교적번호", v("parishioner_no"), "세례명",    v("baptism_nm"))
        r2("세대주",   v("host_nm"),        "관계",      v("relation"))
        r2("주민번호", v("personal_no"),    "축일",      v("baptism_day"))
        r2("등록일",   v("registion_date"), "이전 교구", v("pre_parish_nm"))

        flags = []
        if v("etemal") == "Y":          flags.append("✅ 영원한 교적")
        if v("money_duty_flag") == "Y": flags.append("💰 의무금")
        if v("lazy_flag") == "Y":       flags.append("😴 냉담자")
        if v("alone_flag") == "Y":      flags.append("🏠 독거")
        if v("delete_flag") == "Y":     flags.append("🗑 삭제됨")
        if flags:
            fl = QLabel("  ".join(flags)); fl.setObjectName("mu")
            grid.addWidget(fl, r, 0, 1, 4); r += 1

        sh("📞  연락처")
        r2("집 전화",   v("tel_home"),   "휴대폰",  v("tel_hp"))
        r2("직장 전화", v("tel_office"), "주소",    v("address"))
        r2("상세 주소", v("address_no"))

        sh("✝  세례 정보")
        r2("세례일",    v("baptism_date"),      "세례번호",  v("baptism_no"))
        r2("세례 교구", v("baptism_parish_nm"), "세례 성당", v("baptism_church"))

        sh("🕊  견진 정보")
        r2("견진일",    v("sacrament_date"), "견진 교구", v("sacrament_parish_nm"))

        if v("office_nm") != "—" or v("job_kind") != "—":
            sh("💼  직장")
            r2("직장명", v("office_nm"), "직종", v("job_kind"))

        if v("personal_memo") != "—":
            sh("📝  메모")
            ml = QLabel(v("personal_memo")); ml.setObjectName("fv"); ml.setWordWrap(True)
            grid.addWidget(ml, r, 0, 1, 4); r += 1

        members = self.db.household(v("host_nm"), exclude=pno)
        if members:
            sh(f"👨‍👩‍👧  같은 세대 ({len(members)}명)")
            for m in members:
                txt = (f"  {m['name'].strip()}  ({m['baptism_nm'].strip()})"
                       f"  —  {m['relation'].strip()}  [{m['parishioner_no']}]")
                ml2 = QLabel(txt); ml2.setObjectName("fv")
                grid.addWidget(ml2, r, 0, 1, 4); r += 1

        grid.setRowStretch(r, 1)

        tabs = QTabWidget()
        tabs.addTab(sc, "기본 정보")
        tabs.addTab(SacramentsTab(self.db, pno), "성사 기록")
        tabs.addTab(MoveRecordsTab(self.db, pno), "이동 기록")
        tabs.setCurrentIndex(self._current_tab)
        tabs.currentChanged.connect(lambda i: setattr(self, '_current_tab', i))
        self._l.addWidget(tabs, 1)


# ── Sacrament Add Dialogs ─────────────────────────────────────────────────────

class BaptismForm(QDialog):
    def __init__(self, parent, db, pno, name, on_save=None):
        super().__init__(parent)
        self.db = db; self.pno = pno; self.on_save = on_save
        self.setWindowTitle("세례 기록 추가"); self.resize(600, 240); self.setModal(True)
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        body = QWidget(); body.setObjectName("card")
        g = QGridLayout(body); g.setContentsMargins(16,12,16,12)
        g.setHorizontalSpacing(16); g.setVerticalSpacing(6)
        for col in range(4): g.setColumnStretch(col, 1)
        outer.addWidget(body, 1)

        def add(lbl, w, r, c, span=1):
            g.addWidget(vbox_field(lbl, w, C['card']), r, c, 1, span); return w

        self.no_e    = add("세례 번호",            mk_entry(), 0, 0)
        self.date_e  = add("세례일 (YYYY/MM/DD)",  mk_entry(), 0, 1)
        self.church_e= add("세례 성당",            mk_entry(), 0, 2, 2)
        self.off_e   = add("집전자",               mk_entry(), 1, 0)
        self.god_e   = add("대부/대모",             mk_entry(), 1, 1)
        self.fa_e    = add("부친",                 mk_entry(), 1, 2)
        self.mo_e    = add("모친",                 mk_entry(), 1, 3)

        bb = QWidget(); bb.setObjectName("card"); bb.setFixedHeight(54)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(12,8,12,8); bbl.addStretch()
        cb = mk_btn("취소","btn_muted"); sb = mk_btn("💾  저장","btn_accent")
        cb.clicked.connect(self.reject); sb.clicked.connect(self._save)
        bbl.addWidget(cb); bbl.addWidget(sb); outer.addWidget(bb)

    def _save(self):
        try:
            self.db.create_baptism(dict(
                parishioner_no=self.pno, baptism_no=ge(self.no_e),
                baptism_date=ge(self.date_e), baptism_church=ge(self.church_e),
                officiator_nm=ge(self.off_e), godfather_nm=ge(self.god_e),
                father_nm=ge(self.fa_e), mother_nm=ge(self.mo_e),
            ))
            if self.on_save: self.on_save()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", str(e))


class ConfirmationForm(QDialog):
    def __init__(self, parent, db, pno, name, on_save=None):
        super().__init__(parent)
        self.db = db; self.pno = pno; self.on_save = on_save
        self.setWindowTitle("견진 기록 추가"); self.resize(520, 190); self.setModal(True)
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        body = QWidget(); body.setObjectName("card")
        g = QGridLayout(body); g.setContentsMargins(16,12,16,12)
        g.setHorizontalSpacing(16); g.setVerticalSpacing(6)
        for col in range(4): g.setColumnStretch(col, 1)
        outer.addWidget(body, 1)

        def add(lbl, w, r, c, span=1):
            g.addWidget(vbox_field(lbl, w, C['card']), r, c, 1, span); return w

        self.no_e    = add("견진 번호",            mk_entry(), 0, 0)
        self.date_e  = add("견진일 (YYYY/MM/DD)",  mk_entry(), 0, 1)
        self.church_e= add("견진 성당",            mk_entry(), 0, 2, 2)
        self.off_e   = add("집전자",               mk_entry(), 1, 0, 2)

        bb = QWidget(); bb.setObjectName("card"); bb.setFixedHeight(54)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(12,8,12,8); bbl.addStretch()
        cb = mk_btn("취소","btn_muted"); sb = mk_btn("💾  저장","btn_accent")
        cb.clicked.connect(self.reject); sb.clicked.connect(self._save)
        bbl.addWidget(cb); bbl.addWidget(sb); outer.addWidget(bb)

    def _save(self):
        try:
            self.db.create_sacrament_record(dict(
                parishioner_no=self.pno, sacrament_no=ge(self.no_e),
                sacrament_date=ge(self.date_e), sacrament_parish_church=ge(self.church_e),
                officiator_nm=ge(self.off_e),
            ))
            if self.on_save: self.on_save()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", str(e))


class WeddingForm(QDialog):
    def __init__(self, parent, db, pno, name, on_save=None):
        super().__init__(parent)
        self.db = db; self.pno = pno; self.name = name; self.on_save = on_save
        self.setWindowTitle("혼인 기록 추가"); self.resize(620, 260); self.setModal(True)
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        body = QWidget(); body.setObjectName("card")
        g = QGridLayout(body); g.setContentsMargins(16,12,16,12)
        g.setHorizontalSpacing(16); g.setVerticalSpacing(6)
        for col in range(4): g.setColumnStretch(col, 1)
        outer.addWidget(body, 1)

        def add(lbl, w, r, c, span=1):
            g.addWidget(vbox_field(lbl, w, C['card']), r, c, 1, span); return w

        self.no_e     = add("혼인 번호",            mk_entry(), 0, 0)
        self.date_e   = add("혼인일 (YYYY/MM/DD)",  mk_entry(), 0, 1)
        self.church_e = add("성당",                 mk_entry(), 0, 2, 2)
        self.role_cb  = add("역할",                 mk_combo(["신랑 (m)","신부 (w)"]), 1, 0)
        self.spouse_e = add("배우자 이름",           mk_entry(), 1, 1)
        self.off_e    = add("집전자",               mk_entry(), 1, 2, 2)

        bb = QWidget(); bb.setObjectName("card"); bb.setFixedHeight(54)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(12,8,12,8); bbl.addStretch()
        cb = mk_btn("취소","btn_muted"); sb = mk_btn("💾  저장","btn_accent")
        cb.clicked.connect(self.reject); sb.clicked.connect(self._save)
        bbl.addWidget(cb); bbl.addWidget(sb); outer.addWidget(bb)

    def _save(self):
        try:
            is_groom = ge(self.role_cb).startswith("신랑")
            spouse = ge(self.spouse_e)
            data = dict(
                wedding_no=ge(self.no_e), wedding_date=ge(self.date_e),
                parish_church=ge(self.church_e), officiator_nm=ge(self.off_e),
            )
            if is_groom:
                data["m_parishioner_no"] = self.pno
                data["m_name"]           = self.name
                data["w_name"]           = spouse
            else:
                data["w_parishioner_no"] = self.pno
                data["w_name"]           = self.name
                data["m_name"]           = spouse
            self.db.create_wedding_record(data)
            if self.on_save: self.on_save()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", str(e))


class ConfessionForm(QDialog):
    def __init__(self, parent, db, pno, on_save=None):
        super().__init__(parent)
        self.db = db; self.pno = pno; self.on_save = on_save
        self.setWindowTitle("고해 기록 추가"); self.resize(360, 170); self.setModal(True)
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        body = QWidget(); body.setObjectName("card")
        g = QGridLayout(body); g.setContentsMargins(16,12,16,12)
        g.setHorizontalSpacing(16); g.setVerticalSpacing(8)
        outer.addWidget(body, 1)

        self.year_e  = mk_entry()
        self.spring_cb = mk_check("봄")
        self.fall_cb   = mk_check("가을")
        g.addWidget(vbox_field("연도 (YYYY)", self.year_e, C['card']), 0, 0, 1, 2)
        fw = QWidget(); fw.setObjectName("card")
        fl = QHBoxLayout(fw); fl.setContentsMargins(0,4,0,0)
        fl.addWidget(self.spring_cb); fl.addWidget(self.fall_cb); fl.addStretch()
        g.addWidget(vbox_field("성사 완료", fw, C['card']), 1, 0, 1, 2)

        bb = QWidget(); bb.setObjectName("card"); bb.setFixedHeight(54)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(12,8,12,8); bbl.addStretch()
        cb = mk_btn("취소","btn_muted"); sb = mk_btn("💾  저장","btn_accent")
        cb.clicked.connect(self.reject); sb.clicked.connect(self._save)
        bbl.addWidget(cb); bbl.addWidget(sb); outer.addWidget(bb)

    def _save(self):
        year = ge(self.year_e)
        if not year:
            QMessageBox.warning(self, "오류", "연도를 입력하세요."); return
        try:
            self.db.create_confession_record(dict(
                parishioner_no=self.pno, confession_year=year,
                spring="Y" if self.spring_cb.isChecked() else "N",
                fall="Y" if self.fall_cb.isChecked() else "N",
            ))
            if self.on_save: self.on_save()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", str(e))


class DeathForm(QDialog):
    def __init__(self, parent, db, pno, name, on_save=None):
        super().__init__(parent)
        self.db = db; self.pno = pno; self.on_save = on_save
        self.setWindowTitle("사망 기록 추가"); self.resize(580, 210); self.setModal(True)
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        body = QWidget(); body.setObjectName("card")
        g = QGridLayout(body); g.setContentsMargins(16,12,16,12)
        g.setHorizontalSpacing(16); g.setVerticalSpacing(6)
        for col in range(4): g.setColumnStretch(col, 1)
        outer.addWidget(body, 1)

        def add(lbl, w, r, c, span=1):
            g.addWidget(vbox_field(lbl, w, C['card']), r, c, 1, span); return w

        self.date_e     = add("사망일 (YYYY/MM/DD)",    mk_entry(), 0, 0)
        self.place_e    = add("장소",                   mk_entry(), 0, 1, 2)
        self.off_e      = add("집전자",                 mk_entry(), 0, 3)
        self.sick_e     = add("병자성사일 (YYYY/MM/DD)", mk_entry(), 1, 0)
        self.viat_e     = add("노자성사일 (YYYY/MM/DD)", mk_entry(), 1, 1)

        bb = QWidget(); bb.setObjectName("card"); bb.setFixedHeight(54)
        bbl = QHBoxLayout(bb); bbl.setContentsMargins(12,8,12,8); bbl.addStretch()
        cb = mk_btn("취소","btn_muted"); sb = mk_btn("💾  저장","btn_accent")
        cb.clicked.connect(self.reject); sb.clicked.connect(self._save)
        bbl.addWidget(cb); bbl.addWidget(sb); outer.addWidget(bb)

    def _save(self):
        try:
            self.db.create_death_record(dict(
                parishioner_no=self.pno, death_date=ge(self.date_e),
                place=ge(self.place_e), officiator_nm=ge(self.off_e),
                sickness_date=ge(self.sick_e), viaticum_date=ge(self.viat_e),
            ))
            if self.on_save: self.on_save()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", str(e))


# ── Sacraments Tab ────────────────────────────────────────────────────────────

class SacramentsTab(QWidget):
    def __init__(self, db, pno):
        super().__init__()
        self.db = db; self.pno = pno
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        self._scroll = QScrollArea(); self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.Shape.NoFrame)
        outer.addWidget(self._scroll, 1)
        self._load()

    def reload(self):
        self._load()

    def _load(self):
        body = QWidget(); body.setObjectName("card")
        lay = QVBoxLayout(body); lay.setContentsMargins(12,8,12,16); lay.setSpacing(10)

        p = self.db.get(self.pno)
        name = fv(p, "name") if p else ""

        self._section(lay, "✝  세례", self.db.get_baptism_records(self.pno),
            lambda rec: [
                ("세례번호", fv(rec,"baptism_no")), ("세례일", fv(rec,"baptism_date")),
                ("세례 성당", fv(rec,"baptism_church")), ("집전자", fv(rec,"officiator_nm")),
                ("대부/대모", fv(rec,"godfather_nm")), ("부친", fv(rec,"father_nm")),
                ("모친", fv(rec,"mother_nm")),
            ],
            lambda: BaptismForm(self, self.db, self.pno, name, on_save=self.reload).exec())

        self._section(lay, "🕊  견진", self.db.get_sacrament_records(self.pno),
            lambda rec: [
                ("견진번호", fv(rec,"sacrament_no")), ("견진일", fv(rec,"sacrament_date")),
                ("견진 성당", fv(rec,"sacrament_parish_church")), ("집전자", fv(rec,"officiator_nm")),
            ],
            lambda: ConfirmationForm(self, self.db, self.pno, name, on_save=self.reload).exec())

        self._section(lay, "💒  혼인", self.db.get_wedding_records(self.pno),
            lambda rec: [
                ("혼인번호", fv(rec,"wedding_no")), ("혼인일", fv(rec,"wedding_date")),
                ("성당", fv(rec,"parish_church")), ("집전자", fv(rec,"officiator_nm")),
                ("배우자", fv(rec,"w_name") if fv(rec,"m_parishioner_no").strip()==self.pno.strip()
                           else fv(rec,"m_name")),
            ],
            lambda: WeddingForm(self, self.db, self.pno, name, on_save=self.reload).exec())

        self._section(lay, "📖  고해성사", self.db.get_confession_records(self.pno),
            lambda rec: [
                ("연도", fv(rec,"confession_year")),
                ("봄", "✓" if fv(rec,"spring")=="Y" else "✗"),
                ("가을", "✓" if fv(rec,"fall")=="Y" else "✗"),
            ],
            lambda: ConfessionForm(self, self.db, self.pno, on_save=self.reload).exec())

        self._section(lay, "✟  사망", self.db.get_death_records(self.pno),
            lambda rec: [
                ("사망일", fv(rec,"death_date")), ("장소", fv(rec,"place")),
                ("집전자", fv(rec,"officiator_nm")), ("병자성사", fv(rec,"sickness_date")),
                ("노자성사", fv(rec,"viaticum_date")),
            ],
            lambda: DeathForm(self, self.db, self.pno, name, on_save=self.reload).exec())

        lay.addStretch()
        self._scroll.setWidget(body)

    def _section(self, lay, title, records, fields_fn, on_add):
        hdr_row = QWidget()
        hl = QHBoxLayout(hdr_row); hl.setContentsMargins(0,0,0,0); hl.setSpacing(6)
        hl.addWidget(shdr(title)); hl.addStretch()
        add_btn = mk_btn("+ 추가", "btn_success")
        add_btn.setFixedHeight(26); add_btn.setFixedWidth(68)
        add_btn.clicked.connect(on_add)
        hl.addWidget(add_btn)
        lay.addWidget(hdr_row)

        if not records:
            lbl = QLabel("  기록 없음"); lbl.setObjectName("mu")
            lay.addWidget(lbl)
            return

        for rec in records:
            pairs = [(l, v) for l, v in fields_fn(rec) if v and v not in ("—", "")]
            if not pairs:
                continue
            card = QWidget()
            card.setStyleSheet(f"background:{C['header']};border-radius:4px;")
            g = QGridLayout(card); g.setContentsMargins(10,6,10,6)
            g.setHorizontalSpacing(12); g.setVerticalSpacing(3)
            g.setColumnStretch(1, 2); g.setColumnStretch(3, 2)
            row, col = 0, 0
            for lbl_txt, val_txt in pairs:
                lw = QLabel(lbl_txt); lw.setObjectName("fl")
                vw = QLabel(val_txt); vw.setObjectName("fv"); vw.setWordWrap(True)
                g.addWidget(lw, row, col * 2); g.addWidget(vw, row, col * 2 + 1)
                col += 1
                if col >= 2: col = 0; row += 1
            lay.addWidget(card)


# ── Move Records Tab ─────────────────────────────────────────────────────────

class MoveRecordsTab(QWidget):
    def __init__(self, db, pno):
        super().__init__()
        self.db = db; self.pno = pno
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        outer.addWidget(scroll, 1)

        body = QWidget(); body.setObjectName("card")
        lay = QVBoxLayout(body); lay.setContentsMargins(12,8,12,16); lay.setSpacing(10)

        self._build_movein(lay)
        self._build_moveout(lay)
        lay.addStretch()
        scroll.setWidget(body)

    def _build_movein(self, lay):
        lay.addWidget(shdr("📥  전입 기록"))
        rec = self.db.get_movein_record(self.pno)
        if not rec:
            lbl = QLabel("  기록 없음"); lbl.setObjectName("mu"); lay.addWidget(lbl); return

        area_code = fv(rec, "area_no").strip()
        area_name = AREA_MAP.get(area_code, area_code)
        duty = fv(rec, "duty_money")
        try:
            duty_str = f"{float(duty):,.0f} 원" if duty else ""
        except ValueError:
            duty_str = duty

        pairs = [
            ("전입일",    fv(rec, "movein_date")),
            ("구역",      f"{area_code}  {area_name}" if area_name else area_code),
            ("이전 교구", fv(rec, "pre_parish_nm")),
            ("이전 성당", fv(rec, "pre_parish_church")),
            ("의무금",    duty_str),
        ]
        lay.addWidget(self._record_card(pairs))

    def _build_moveout(self, lay):
        lay.addWidget(shdr("📤  전출 기록"))
        records = self.db.get_moveout_records(self.pno)
        if not records:
            lbl = QLabel("  기록 없음"); lbl.setObjectName("mu"); lay.addWidget(lbl); return

        for rec in records:
            duty = fv(rec, "duty_money")
            try:
                duty_str = f"{float(duty):,.0f} 원" if duty else ""
            except ValueError:
                duty_str = duty

            pairs = [
                ("전출일",  fv(rec, "moveout_date")),
                ("새 교구", fv(rec, "new_parish_nm")),
                ("새 성당", fv(rec, "new_parish_church")),
                ("의무금",  duty_str),
            ]
            lay.addWidget(self._record_card(pairs))

    def _record_card(self, pairs):
        pairs = [(l, v) for l, v in pairs if v and v not in ("—", "")]
        card = QWidget()
        card.setStyleSheet(f"background:{C['header']};border-radius:4px;")
        g = QGridLayout(card); g.setContentsMargins(10,6,10,6)
        g.setHorizontalSpacing(12); g.setVerticalSpacing(3)
        g.setColumnStretch(1, 2); g.setColumnStretch(3, 2)
        row, col = 0, 0
        for lbl_txt, val_txt in pairs:
            lw = QLabel(lbl_txt); lw.setObjectName("fl")
            vw = QLabel(val_txt); vw.setObjectName("fv"); vw.setWordWrap(True)
            g.addWidget(lw, row, col*2); g.addWidget(vw, row, col*2+1)
            col += 1
            if col >= 2: col = 0; row += 1
        return card


# ── Export Dialog ─────────────────────────────────────────────────────────────

class ExportDialog(QDialog):
    HEADERS = ["교적번호", "이름", "세례명", "관계", "세대주", "전화번호", "주소", "등록일"]
    FIELDS  = ["parishioner_no", "name", "baptism_nm", "relation",
               "host_nm", "tel_home", "address", "registion_date"]

    def __init__(self, parent, rows):
        super().__init__(parent)
        self.rows = rows
        self.setWindowTitle("목록 내보내기")
        self.resize(360, 150)
        self.setModal(True)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(12)

        info = QLabel(f"현재 표시된 교적 {len(rows)}명을 내보냅니다.")
        info.setObjectName("fv")
        lay.addWidget(info)

        btn_row = QHBoxLayout(); btn_row.setSpacing(8)
        csv_btn  = mk_btn("CSV (.csv)",    "btn_accent")
        xlsx_btn = mk_btn("Excel (.xlsx)", "btn_success")
        csv_btn.clicked.connect(self._export_csv)
        xlsx_btn.clicked.connect(self._export_xlsx)
        btn_row.addWidget(csv_btn); btn_row.addWidget(xlsx_btn)
        lay.addLayout(btn_row)

        cancel = mk_btn("취소", "btn_muted")
        cancel.clicked.connect(self.reject)
        lay.addWidget(cancel)

    def _values(self):
        return [[str(r[f] or "").strip() for f in self.FIELDS] for r in self.rows]

    def _export_csv(self):
        import csv
        path, _ = QFileDialog.getSaveFileName(
            self, "CSV로 저장", "교적목록.csv", "CSV Files (*.csv)")
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(self.HEADERS)
                w.writerows(self._values())
            QMessageBox.information(self, "완료", f"저장 완료:\n{path}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))

    def _export_xlsx(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(
                self, "오류",
                "openpyxl이 설치되지 않았습니다.\n터미널에서 실행하세요: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel로 저장", "교적목록.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "교적 목록"

            ws.append(self.HEADERS)
            header_fill = PatternFill("solid", fgColor="2D4A6A")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 20

            for row in self._values():
                ws.append(row)

            for col in ws.columns:
                width = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(width + 4, 40)

            wb.save(path)
            QMessageBox.information(self, "완료", f"저장 완료:\n{path}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))


# ── List View ─────────────────────────────────────────────────────────────────

class ListView(QWidget):
    row_selected = pyqtSignal(str)

    def __init__(self, db):
        super().__init__()
        self.db = db
        self._rows = []
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        tb = QWidget(); tb.setObjectName("toolbar"); tb.setFixedHeight(52)
        tbl = QHBoxLayout(tb); tbl.setContentsMargins(14, 8, 14, 8)
        tl = QLabel("교적 목록")
        tl.setStyleSheet(f"font-size:15px;font-weight:bold;color:{C['text']};background:transparent;")
        self.export_btn = mk_btn("⬇  내보내기", "btn_muted")
        self.add_btn    = mk_btn("+ 새 교적",   "btn_accent")
        self.export_btn.clicked.connect(lambda: ExportDialog(self, self._rows).exec())
        tbl.addWidget(tl); tbl.addStretch()
        tbl.addWidget(self.export_btn); tbl.addWidget(self.add_btn)
        lay.addWidget(tb)

        fb = QWidget(); fb.setObjectName("filterbar"); fb.setFixedHeight(48)
        fbl = QHBoxLayout(fb); fbl.setContentsMargins(10, 6, 10, 6); fbl.setSpacing(8)
        self.q_le = QLineEdit()
        self.q_le.setPlaceholderText("🔍  이름 / 세례명 / 교적번호 / 세대주")
        self.q_le.setFixedWidth(280)
        self.area_cb = QComboBox()
        self.area_cb.addItem("전체 구역")
        self.area_cb.addItems(AREA_DISP)
        self.area_cb.setFixedWidth(180)
        self.del_cb  = QCheckBox("삭제된 교적 포함")
        self.cnt_lbl = QLabel(""); self.cnt_lbl.setObjectName("mu")
        fbl.addWidget(self.q_le); fbl.addWidget(self.area_cb); fbl.addWidget(self.del_cb)
        fbl.addStretch(); fbl.addWidget(self.cnt_lbl)
        lay.addWidget(fb)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["교적번호", "이름", "세례명", "관계", "세대주", "전화번호"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        for i, w in enumerate([130, 100, 120, 70, 100, 120]):
            self.table.setColumnWidth(i, w)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(f"alternate-background-color:{C['row_alt']};")
        lay.addWidget(self.table, 1)

        self.q_le.textChanged.connect(self.load)
        self.area_cb.currentIndexChanged.connect(self.load)
        self.del_cb.stateChanged.connect(self.load)
        self.table.itemSelectionChanged.connect(self._sel)

    def load(self):
        q    = self.q_le.text()
        area = self.area_cb.currentText()
        area = area.split()[0] if area and area != "전체 구역" else ""
        rows = self.db.search(q, area, self.del_cb.isChecked())
        self._rows = rows
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            vals = [r["parishioner_no"], r["name"] or "", r["baptism_nm"] or "",
                    r["relation"] or "", r["host_nm"] or "", r["tel_home"] or ""]
            is_del = str(r["delete_flag"]).strip() == "Y"
            is_lap = str(r["lazy_flag"]).strip() == "Y"
            for j, val in enumerate(vals):
                it = QTableWidgetItem(val)
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if is_del:
                    it.setForeground(QColor("#AAAAAA"))
                elif is_lap:
                    it.setForeground(QColor(C["orange"]))
                self.table.setItem(i, j, it)
        self.cnt_lbl.setText(f"총 {len(rows)}명")

    def _sel(self):
        row = self.table.currentRow()
        if row >= 0:
            pno = self.table.item(row, 0)
            if pno:
                self.row_selected.emit(pno.text())


# ── Stats View ────────────────────────────────────────────────────────────────

class StatsView(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 16)
        title = QLabel("현황 통계")
        title.setStyleSheet(f"font-size:18px;font-weight:bold;color:{C['text']};")
        lay.addWidget(title)

        s = self.db.stats()
        cr = QHBoxLayout(); cr.setSpacing(12)
        for label, val, col in [
            ("👥  활성 교인",  s["active"],   C["accent"]),
            ("😴  냉담자",     s["lapsed"],   C["orange"]),
            ("✝  세례 기록",  s["baptisms"], C["success"]),
            ("💒  혼인 기록",  s["weddings"], "#8E44AD"),
        ]:
            card = QWidget(); card.setObjectName("stat_card"); card.setFixedSize(170, 95)
            cl = QVBoxLayout(card); cl.setContentsMargins(8, 8, 8, 8)
            nl = QLabel(str(val)); nl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            nl.setStyleSheet(f"font-size:34px;font-weight:bold;color:{col};background:transparent;")
            ll = QLabel(label); ll.setAlignment(Qt.AlignmentFlag.AlignCenter)
            ll.setStyleSheet(f"font-size:11px;color:{C['muted']};background:transparent;")
            cl.addWidget(nl); cl.addWidget(ll); cr.addWidget(card)
        cr.addStretch()
        lay.addLayout(cr)
        lay.addSpacing(16)

        al = QLabel("구역별 교인 현황")
        al.setStyleSheet(f"font-size:14px;font-weight:bold;color:{C['text']};")
        lay.addWidget(al)
        lay.addSpacing(6)

        for i, row in enumerate(s["areas"]):
            name = AREA_MAP.get(row["area"], row["area"])
            rf = QWidget()
            rf.setStyleSheet(
                f"background:{C['header'] if i % 2 == 0 else C['card']};border-radius:4px;"
            )
            rl = QHBoxLayout(rf); rl.setContentsMargins(12, 4, 12, 4)
            la = QLabel(f"{row['area']}  {name}")
            la.setStyleSheet("font-size:12px;background:transparent;")
            lc = QLabel(f"{row['cnt']}명")
            lc.setStyleSheet(f"font-size:12px;font-weight:bold;color:{C['accent']};background:transparent;")
            rl.addWidget(la); rl.addStretch(); rl.addWidget(lc)
            lay.addWidget(rf)
        lay.addStretch()


# ── Main Window ───────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setWindowTitle(APP_TITLE)
        self.resize(1300, 820)
        self.setMinimumSize(960, 600)

        root = QWidget(); self.setCentralWidget(root)
        rl = QHBoxLayout(root); rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(0)

        # ── Sidebar ──────────────────────────────────────────────────────────
        sb = QWidget(); sb.setObjectName("sidebar"); sb.setFixedWidth(196)
        sbl = QVBoxLayout(sb); sbl.setContentsMargins(8, 16, 8, 12); sbl.setSpacing(4)
        for txt, style in [
            ("✝",                      "font-size:28px;color:white;background:transparent;"),
            ("교적 관리",               "font-size:14px;font-weight:bold;color:white;background:transparent;"),
            ("Boston Korean Catholic",  "font-size:9px;color:#8EAFD4;background:transparent;"),
        ]:
            l = QLabel(txt)
            l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            l.setStyleSheet(style)
            sbl.addWidget(l)
        sbl.addSpacing(16)

        self.nav_list  = QPushButton("👥  교적 목록")
        self.nav_stats = QPushButton("📊  현황 통계")
        for nb in [self.nav_list, self.nav_stats]:
            nb.setObjectName("nav_btn")
            nb.setCursor(Qt.CursorShape.PointingHandCursor)
            sbl.addWidget(nb)

        sbl.addSpacing(12)
        self._count_lbl = QLabel()
        self._count_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._count_lbl.setStyleSheet("font-size:11px;color:#8EAFD4;background:transparent;")
        sbl.addWidget(self._count_lbl)

        sbl.addStretch()
        db_lbl = QLabel(os.path.basename(DB_PATH))
        db_lbl.setStyleSheet("font-size:8px;color:#5A7FA8;background:transparent;")
        db_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        db_lbl.setWordWrap(True)
        sbl.addWidget(db_lbl)
        rl.addWidget(sb)

        # ── Main stack ───────────────────────────────────────────────────────
        self.stack = QStackedWidget(); rl.addWidget(self.stack, 1)

        self.list_view   = ListView(db)
        self.detail_view = DetailPanel(db)
        sp = QSplitter(Qt.Orientation.Horizontal)
        sp.addWidget(self.list_view); sp.addWidget(self.detail_view)
        sp.setSizes([620, 380]); sp.setHandleWidth(1)
        lp = QWidget()
        lpl = QVBoxLayout(lp); lpl.setContentsMargins(0, 0, 0, 0); lpl.setSpacing(0)
        lpl.addWidget(sp)
        self.stack.addWidget(lp)

        self.stats_view = StatsView(db)
        self.stack.addWidget(self.stats_view)

        # ── Signals ──────────────────────────────────────────────────────────
        self.nav_list.clicked.connect(lambda: self.stack.setCurrentIndex(0))
        self.nav_stats.clicked.connect(lambda: self.stack.setCurrentIndex(1))
        self.list_view.add_btn.clicked.connect(self._add)
        self.list_view.row_selected.connect(self.detail_view.load)
        self.detail_view.edit_sig.connect(self._edit)
        self.detail_view.delete_sig.connect(self._delete)
        self.detail_view.perm_delete_sig.connect(self._perm_delete)
        self.detail_view.add_sig.connect(self._add_member)

        # ── Keyboard shortcuts ────────────────────────────────────────────────
        QShortcut(QKeySequence("Ctrl+N"), self).activated.connect(self._add)
        QShortcut(QKeySequence("Ctrl+F"), self).activated.connect(self._focus_search)
        QShortcut(QKeySequence("Escape"), self).activated.connect(self._clear_search)

        # ── Restore geometry then load data ───────────────────────────────────
        settings = QSettings("BostonKoreanCatholic", "ChurchRegistry")
        geom = settings.value("window/geometry")
        if geom:
            self.restoreGeometry(geom)

        self._reload()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _reload(self):
        self.list_view.load()
        self._refresh_count()

    def _refresh_count(self):
        count = self.db.stats()["active"]
        self._count_lbl.setText(f"활성 교인  {count}명")

    def _focus_search(self):
        self.stack.setCurrentIndex(0)
        self.list_view.q_le.setFocus()

    def _clear_search(self):
        self.stack.setCurrentIndex(0)
        self.list_view.q_le.clear()

    def closeEvent(self, event):
        QSettings("BostonKoreanCatholic", "ChurchRegistry").setValue(
            "window/geometry", self.saveGeometry())
        super().closeEvent(event)

    # ── CRUD actions ──────────────────────────────────────────────────────────

    def _add(self):
        ParishionerForm(self, self.db, on_save=self._reload).exec()

    def _edit(self, pno):
        def refresh():
            self._reload()
            self.detail_view.load(pno)
        ParishionerForm(self, self.db, pno=pno, on_save=refresh).exec()

    def _add_member(self, host, ref_pno):
        area = ref_pno.split("-")[0] if "-" in ref_pno else "00112"
        def refresh():
            self._reload()
            self.detail_view.load(ref_pno)
        ParishionerForm(self, self.db, prefill_host=host, prefill_area=area, on_save=refresh).exec()

    def _perm_delete(self, pno):
        msg = (f"교적을 영구적으로 삭제하시겠습니까?\n{pno}\n\n"
               "이 작업은 되돌릴 수 없습니다.")
        if QMessageBox.question(self, "영구 삭제 확인", msg) == QMessageBox.StandardButton.Yes:
            self.db.hard_delete(pno)
            self._reload()
            self.detail_view.show_empty()

    def _delete(self, pno, is_deleted):
        if is_deleted:
            if QMessageBox.question(self, "복원 확인", f"교적을 복원하시겠습니까?\n{pno}") == QMessageBox.StandardButton.Yes:
                self.db.restore(pno)
                self._reload()
                self.detail_view.load(pno)
        else:
            if QMessageBox.question(self, "삭제 확인", f"교적을 삭제하시겠습니까?\n{pno}\n\n나중에 복원할 수 있습니다.") == QMessageBox.StandardButton.Yes:
                self.db.soft_delete(pno)
                self._reload()
                self.detail_view.show_empty()
